import os
import openpyxl
import pandas as pd

from datetime import datetime
from pathlib import Path
from openpyxl.styles import Border, Side
from Utils import Utils


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    pass


def getExcelData(fPath):
    fileName = Path(fPath).stem

    print("Reading: " + fPath + "......")
    print("==============================================================")

    util = Utils()
    df = pd.read_excel(fPath, sheet_name='Eligible Population',
                       header=1, dtype={'PhilHealth_ID*': str, 'Contact_number_of_employer*': str,
                                        'Contact_No.*': str, 'Age': str, 'Company': str}, na_filter=False)

    groups = df.groupby('Company')

    for comp, records in groups:
        outputFileName = comp + "_AZ"
        records = records.astype(str)

        # get num rows
        numrows = len(records.index)

        print(comp + "  has " + str(numrows) + " records", end='')
        # print(comp + " --> has " + str(numrows) + " records")

        templateFile = util.duplicateTemplateLTGC(templateFilePath, outPath, outputFileName)

        theFile = openpyxl.load_workbook(templateFile)
        currentSheet = theFile["Eligible Population"]
        util.addingDataValidation(currentSheet, numrows)

        set_border(currentSheet, "A3:BW" + str(numrows + 2))

        theFile.save(templateFile)

        writer = pd.ExcelWriter(templateFile, engine='openpyxl', mode='a')
        writer.book = openpyxl.load_workbook(templateFile)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        records.to_excel(writer, sheet_name="Eligible Population", startrow=2, header=False, index=False)
        writer.save()

        print("...Done...")

    pass


if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', None)

    today = datetime.today()
    dateTime = today.strftime("%m_%d_%y_%H%M%S")

    dirPath = r"/Users/Ran/Documents/Vaccine/LTGSplit_sub-comp"

    inPath = os.path.join(dirPath, "in")
    outPath = os.path.join(dirPath, "out/emp")
    templateFilePath = os.path.join(dirPath, "template/LTGC_CEIRMasterlist_ExtraCols.xlsx")

    print("==============================================================")
    print("Running Scpirt: Split Sub Companies......")
    print("==============================================================")

    # Get all filenames from folder and convert to list
    arrFilenames = os.listdir(inPath)

    for filename in arrFilenames:
        keyCeirMaster = filename.split("_")[0]

        if not filename == ".DS_Store":
            FilePath = os.path.join(inPath, filename)

            getExcelData(FilePath)

